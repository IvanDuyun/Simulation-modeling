

from PyQt5.QtWidgets import QWidget, QApplication, QPushButton, QLabel, QHBoxLayout, QVBoxLayout, QLineEdit, QComboBox
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from PyQt5.QtWidgets import QTableWidget, QTableWidgetItem
from random import choice
from pylab import *
from Evm import Evm
from Task import Task
import random_values
from math import *
import openpyxl

class MainWindow(QWidget):
    def __init__(self):
        # Вызвается метод суперкласса
        super().__init__()
        # Запускается метод, в котором прописан интерфейс
        self.interface()


    def interface(self):
        '''Создаёт интерфейс'''
        #self.setWindowFlags(Qt.MSWindowsFixedSizeDialogHint)
        self.resize(1000, 950)
        #self.setWindowIcon(QIcon('icon.jpg'))
        self.setWindowTitle('SMO')

        self.shell = QVBoxLayout()

        self.main_box = QHBoxLayout()
        self.shell.addLayout(self.main_box)

        self.modeling_setting_box = QVBoxLayout()
        self.main_box.addLayout(self.modeling_setting_box)

        self.det_model_info = QLabel('Настройки детерминированной модели СМО')
        self.det_model_info.setMaximumHeight(10)
        self.modeling_setting_box.addWidget(self.det_model_info)

        self.size_ozu_box = QHBoxLayout()
        self.size_ozu_info = QLabel('Емкость ОЗУ:')
        self.size_ozu_cnt = QLineEdit('2')
        self.modeling_setting_box.addLayout(self.size_ozu_box)
        self.size_ozu_box.addWidget(self.size_ozu_info)
        self.size_ozu_box.addWidget(self.size_ozu_cnt)

        self.time_evm1_task_box = QHBoxLayout()
        self.time_evm1_task_info = QLabel('Время обработки задания ЭВМ1:')
        self.time_evm1_task_cnt = QLineEdit('30')
        self.time_evm1_task_box.addWidget(self.time_evm1_task_info)
        self.time_evm1_task_box.addWidget(self.time_evm1_task_cnt)
        self.modeling_setting_box.addLayout(self.time_evm1_task_box)

        self.time_evm2_task_box = QHBoxLayout()
        self.time_evm2_task_info = QLabel('Время обработки задания ЭВМ2:')
        self.time_evm2_task_cnt = QLineEdit('14')
        self.time_evm2_task_box.addWidget(self.time_evm2_task_info)
        self.time_evm2_task_box.addWidget(self.time_evm2_task_cnt)
        self.modeling_setting_box.addLayout(self.time_evm2_task_box)

        self.time_evm3_task_box = QHBoxLayout()
        self.time_evm3_task_info = QLabel('Время обработки задания ЭВМ3:')
        self.time_evm3_task_cnt = QLineEdit('16')
        self.time_evm3_task_box.addWidget(self.time_evm3_task_info)
        self.time_evm3_task_box.addWidget(self.time_evm3_task_cnt)
        self.modeling_setting_box.addLayout(self.time_evm3_task_box)

        self.time_background_task_box = QHBoxLayout()
        self.time_background_task_info = QLabel('Время обработки фоновой задачи:')
        self.time_background_task_cnt = QLineEdit('60')
        self.time_background_task_box.addWidget(self.time_background_task_info)
        self.time_background_task_box.addWidget(self.time_background_task_cnt)
        self.modeling_setting_box.addLayout(self.time_background_task_box)

        self.time_between_tasks_box = QHBoxLayout()
        self.time_between_tasks_info = QLabel('Время между появлением задач:')
        self.time_between_tasks_cnt = QLineEdit('30')
        self.time_between_tasks_box.addWidget(self.time_between_tasks_info)
        self.time_between_tasks_box.addWidget(self.time_between_tasks_cnt)
        self.modeling_setting_box.addLayout(self.time_between_tasks_box)

        self.modeling_time_box = QHBoxLayout()
        self.modeling_setting_box.addLayout(self.modeling_time_box)
        self.modeling_time_info = QLabel('Время моделирования:')
        self.modeling_time_box.addWidget(self.modeling_time_info)
        self.modeling_time_cnt = QLineEdit('3600')
        self.modeling_time_box.addWidget(self.modeling_time_cnt)

        self.imitation_model_info = QLabel('Настройки имитационной модели СМО')
        self.imitation_model_info.setMaximumHeight(10)
        self.modeling_setting_box.addWidget(self.imitation_model_info)

        self.prmtrs_evm1_info = QLabel('Параметры для времени обработки задания ЭВМ1')
        self.prmtrs_evm1_info.setMaximumHeight(20)
        self.modeling_setting_box.addWidget(self.prmtrs_evm1_info)
        self.prmtrs_evm1_Hbox = QHBoxLayout()
        self.modeling_setting_box.addLayout(self.prmtrs_evm1_Hbox)
        self.prmtrs_evm1_m_info = QLabel('m:')
        self.prmtrs_evm1_m_cnt = QLineEdit('30')
        self.prmtrs_evm1_sigma_info = QLabel('sigma:')
        self.prmtrs_evm1_sigma_cnt = QLineEdit('3')
        self.prmtrs_evm1_Hbox.addWidget(self.prmtrs_evm1_m_info)
        self.prmtrs_evm1_Hbox.addWidget(self.prmtrs_evm1_m_cnt)
        self.prmtrs_evm1_Hbox.addWidget(self.prmtrs_evm1_sigma_info)
        self.prmtrs_evm1_Hbox.addWidget(self.prmtrs_evm1_sigma_cnt)

        self.prmtrs_evm2_info = QLabel('Параметры для времени обработки задания ЭВМ2')
        self.prmtrs_evm2_info.setMaximumHeight(20)
        self.modeling_setting_box.addWidget(self.prmtrs_evm2_info)
        self.prmtrs_evm2_Hbox = QHBoxLayout()
        self.modeling_setting_box.addLayout(self.prmtrs_evm2_Hbox)
        self.prmtrs_evm2_m_info = QLabel('m:')
        self.prmtrs_evm2_m_cnt = QLineEdit('14')
        self.prmtrs_evm2_sigma_info = QLabel('sigma:')
        self.prmtrs_evm2_sigma_cnt = QLineEdit('2')
        self.prmtrs_evm2_Hbox.addWidget(self.prmtrs_evm2_m_info)
        self.prmtrs_evm2_Hbox.addWidget(self.prmtrs_evm2_m_cnt)
        self.prmtrs_evm2_Hbox.addWidget(self.prmtrs_evm2_sigma_info)
        self.prmtrs_evm2_Hbox.addWidget(self.prmtrs_evm2_sigma_cnt)

        self.prmtrs_evm3_info = QLabel('Параметры для времени обработки задания ЭВМ3')
        self.prmtrs_evm3_info.setMaximumHeight(20)
        self.modeling_setting_box.addWidget(self.prmtrs_evm3_info)
        self.prmtrs_evm3_Hbox = QHBoxLayout()
        self.modeling_setting_box.addLayout(self.prmtrs_evm3_Hbox)
        self.prmtrs_evm3_m_info = QLabel('m:')
        self.prmtrs_evm3_m_cnt = QLineEdit('16')
        self.prmtrs_evm3_sigma_info = QLabel('sigma:')
        self.prmtrs_evm3_sigma_cnt = QLineEdit('2')
        self.prmtrs_evm3_Hbox.addWidget(self.prmtrs_evm3_m_info)
        self.prmtrs_evm3_Hbox.addWidget(self.prmtrs_evm3_m_cnt)
        self.prmtrs_evm3_Hbox.addWidget(self.prmtrs_evm3_sigma_info)
        self.prmtrs_evm3_Hbox.addWidget(self.prmtrs_evm3_sigma_cnt)

        self.time_between_tasks_random_info = QLabel('Параметр lambda для времени между появлением заданий')
        self.modeling_setting_box.addWidget(self.time_between_tasks_random_info)
        self.time_between_tasks_lambda_box = QHBoxLayout()
        self.modeling_setting_box.addLayout(self.time_between_tasks_lambda_box)
        self.time_between_tasks_lambda_info = QLabel('lambda:')
        self.time_between_tasks_lambda_box.addWidget(self.time_between_tasks_lambda_info)
        self.time_between_tasks_lambda_cnt = QLineEdit('0.03')
        self.time_between_tasks_lambda_box.addWidget(self.time_between_tasks_lambda_cnt)

        self.solution_task_random_info = QLabel('Интервал разброса времён решения фонового задания')
        self.modeling_setting_box.addWidget(self.solution_task_random_info)
        self.solution_task_random_box = QHBoxLayout()
        self.solution_task_random_min_info = QLabel('min:')
        self.solution_task_random_box.addWidget(self.solution_task_random_min_info)
        self.solution_task_random_min_cnt = QLineEdit('40')
        self.solution_task_random_box.addWidget(self.solution_task_random_min_cnt)
        self.solution_task_random_max_info = QLabel('max:')
        self.solution_task_random_box.addWidget(self.solution_task_random_max_info)
        self.solution_task_random_max_cnt = QLineEdit('80')
        self.solution_task_random_box.addWidget(self.solution_task_random_max_cnt)
        self.modeling_setting_box.addLayout(self.solution_task_random_box)

        self.model_type_switch = QComboBox(self)
        self.model_type_switch.addItems(['Детерминированная модель', 'Имитационная модель'])
        self.modeling_setting_box.addWidget(self.model_type_switch)

        self.statistics_box = QVBoxLayout()
        self.modeling_setting_box.addLayout(self.statistics_box)
        self.statistics_workload_evm23 = QLabel('Степень загруженности ЭВМ2,3: ')
        self.statistics_box.addWidget(self.statistics_workload_evm23)
        self.statistics_workload_ozu = QLabel('Степень загруженности ОЗУ: ')
        self.statistics_box.addWidget(self.statistics_workload_ozu)
        self.statistics_p_del_task = QLabel('Вероятность уничтожения задания: ')
        self.statistics_box.addWidget(self.statistics_p_del_task)

        self.button_modeling = QPushButton('Моделирование')
        self.modeling_setting_box.addWidget(self.button_modeling)
        self.button_modeling.clicked.connect(self.computing_device_modeling)

        self.random_test_box = QVBoxLayout()

        self.random_test_settings_box = QHBoxLayout()
        self.random_test_m_info = QLabel('m:')
        self.random_test_m_cnt = QLineEdit('30')
        self.random_test_sigma_info = QLabel('sigma:')
        self.random_test_sigma_cnt = QLineEdit('3')
        self.random_test_lambda_info = QLabel('lambda:')
        self.random_test_lambda_cnt = QLineEdit('0.03')
        self.random_test_settings_box.addWidget(self.random_test_m_info)
        self.random_test_settings_box.addWidget(self.random_test_m_cnt)
        self.random_test_settings_box.addWidget(self.random_test_sigma_info)
        self.random_test_settings_box.addWidget(self.random_test_sigma_cnt)
        self.random_test_settings_box.addWidget(self.random_test_lambda_info)
        self.random_test_settings_box.addWidget(self.random_test_lambda_cnt)
        self.random_test_box.addLayout(self.random_test_settings_box)

        self.graph_switch_box = QHBoxLayout()
        self.distribution_switch = QComboBox()
        self.distribution_switch.addItems(['Нормальное распределение', 'Экспоненциальное распределение'])
        self.graph_switch_box.addWidget(self.distribution_switch)
        self.random_method_switch = QComboBox()
        self.random_method_switch.addItems(
            ['Метод серединных квадратов', 'Метод иррационального числа', 'Конгруэнтный метод'])
        self.graph_switch_box.addWidget(self.random_method_switch)
        self.random_test_box.addLayout(self.graph_switch_box)

        self.figure_1 = plt.figure(facecolor='0.945')
        self.chart_1 = FigureCanvas(self.figure_1)
        self.random_test_box.addWidget(self.chart_1)

        self.SKO = QLabel('Среднеквадратичное отклонение:')
        self.random_test_box.addWidget(self.SKO)

        self.figure_2 = plt.figure(facecolor='0.945')
        self.chart_2 = FigureCanvas(self.figure_2)
        self.random_test_box.addWidget(self.chart_2)

        self.kolmogorov_hypothesis = QLabel('Критерий согласия Колмогорова: ')
        self.random_test_box.addWidget(self.kolmogorov_hypothesis)

        self.pirson_hypothesis = QLabel('Критерий согласия Пирсона: ')
        self.random_test_box.addWidget(self.pirson_hypothesis)

        self.korel_moment = QLabel('Корелляционный метод: ')
        self.random_test_box.addWidget(self.korel_moment)

        self.button_test_random = QPushButton('Тест')
        self.random_test_box.addWidget(self.button_test_random)

        self.main_box.addLayout(self.random_test_box)

        self.disturbing_proc_box = QVBoxLayout()

        self.disturb_data_Hbox = QHBoxLayout()
        self.ymin_info = QLabel('ymin:')
        self.disturb_data_Hbox.addWidget(self.ymin_info)
        self.ymin_cnt = QLineEdit('40')
        self.disturb_data_Hbox.addWidget(self.ymin_cnt)
        self.ymax_info = QLabel('ymax:')
        self.disturb_data_Hbox.addWidget(self.ymax_info)
        self.ymax_cnt = QLineEdit('80')
        self.disturb_data_Hbox.addWidget(self.ymax_cnt)
        self.disturbing_proc_box.addLayout(self.disturb_data_Hbox)

        self.figure_3 = plt.figure(facecolor='0.945')
        self.chart_3 = FigureCanvas(self.figure_3)
        self.disturbing_proc_box.addWidget(self.chart_3)

        self.figure_4 = plt.figure(facecolor='0.945')
        self.chart_4 = FigureCanvas(self.figure_4)
        self.disturbing_proc_box.addWidget(self.chart_4)

        self.figure_5 = plt.figure(facecolor='0.945')
        self.chart_5 = FigureCanvas(self.figure_5)
        self.disturbing_proc_box.addWidget(self.chart_5)

        self.figure_6 = plt.figure(facecolor='0.945')
        self.chart_6 = FigureCanvas(self.figure_6)
        self.disturbing_proc_box.addWidget(self.chart_6)

        self.figure_7 = plt.figure(facecolor='0.945')
        self.chart_7 = FigureCanvas(self.figure_7)
        self.disturbing_proc_box.addWidget(self.chart_7)

        self.smirnov_info = QLabel('Критерий согласия Смирнова')
        self.disturbing_proc_box.addWidget(self.smirnov_info)
        self.smirnov_info_output = QLabel('Cтатистическая гипотеза верна с вероятностью не менее, чем ')
        self.disturbing_proc_box.addWidget(self.smirnov_info_output)

        self.stud_info = QLabel('Критерий согласия Стьюдента')
        self.disturbing_proc_box.addWidget(self.stud_info)
        self.stud_info_output = QLabel('Cтатистическая гипотеза верна с вероятностью не менее, чем ')
        self.disturbing_proc_box.addWidget(self.stud_info_output)

        self.fisher_info = QLabel('Критерий согласия Фишера')
        self.disturbing_proc_box.addWidget(self.fisher_info)
        self.fisher_info_output = QLabel('Cтатистическая гипотеза верна с вероятностью не менее, чем ')
        self.disturbing_proc_box.addWidget(self.fisher_info_output)

        self.disturbing_button = QPushButton('Тест')
        self.disturbing_button.clicked.connect(self.disturbing_proc_generate)
        self.disturbing_proc_box.addWidget(self.disturbing_button)

        self.disturbing_proc_generate()

        self.main_box.addLayout(self.disturbing_proc_box)

        self.button_test_random.clicked.connect(self.test_distribution)

        self.tableWidget = QTableWidget()
        self.tableWidget.setMinimumHeight(200)
        self.tableWidget.setColumnCount(3)
        self.shell.addWidget(self.tableWidget)

        self.setLayout(self.shell)

        self.test_distribution()

        self.show()

    def read_start_values(self):
        self.size_ozu = int(self.size_ozu_cnt.text())
        self.time_be4_cmplt_evm2 = int(self.time_evm2_task_cnt.text())
        self.time_background_task = int(self.time_background_task_cnt.text())
        self.time_be4_cmplt_evm3 = int(self.time_evm3_task_cnt.text())
        self.tm_between_newtasks = int(self.time_between_tasks_cnt.text())
        self.time_be4_cmplt_evm1 = int(self.time_evm1_task_cnt.text())
        self.global_tm = int(self.modeling_time_cnt.text())
        self.evm1_m = float(self.prmtrs_evm1_m_cnt.text())
        self.evm1_sigma = float(self.prmtrs_evm1_sigma_cnt.text())
        self.evm2_m = float(self.prmtrs_evm2_m_cnt.text())
        self.evm2_sigma = float(self.prmtrs_evm2_sigma_cnt.text())
        self.evm3_m = float(self.prmtrs_evm3_m_cnt.text())
        self.evm3_sigma = float(self.prmtrs_evm3_sigma_cnt.text())
        self.tm_be4_newtask_lambda = float(self.time_between_tasks_lambda_cnt.text())
        self.sln_task_tm_rnd_min = int(self.solution_task_random_min_cnt.text())
        self.sln_task_tm_rnd_max = int(self.solution_task_random_max_cnt.text())
        self.ymin = int(self.ymin_cnt.text())
        self.ymax = int(self.ymax_cnt.text())

    def max_list_el(self, lst):
        max = lst[0]
        for el in lst:
            if el > max:
                max = el
        return max

    def min_list_el(self, lst):
        min = lst[0]
        for el in lst:
            if el < min:
                min = el
        return min

    def build_graph_gistogramm(self, switch, coord_x=None, coord_y=None, color=None):
        if switch == 'create':
            self.figure_1.clear()
            self.bar_chart = self.figure_1.add_subplot(111)
            self.bar_chart.grid(True)
        elif switch == 'build':
            self.bar_chart.plot(coord_x, coord_y, color)
            self.chart_1.draw()

    def build_graph_acceptance_criterion(self, switch, coord_x=None, coord_y=None, color=None):
        if switch == 'create':
            self.figure_2.clear()
            self.acceptance_criterion_chart = self.figure_2.add_subplot(111)
            self.acceptance_criterion_chart.grid(True)
        elif switch == 'build':
            self.acceptance_criterion_chart.plot(coord_x, coord_y, color)
            self.chart_2.draw()

    def build_graph(self, switch, figure, chart, coord_x=None, coord_y=None, color=None):
        if switch == 'create':
            figure.clear()
            self.correl_criterion_chart = figure.add_subplot(111)
            self.correl_criterion_chart.grid(True)
        elif switch == 'build':
            self.correl_criterion_chart.plot(coord_x, coord_y, color)
            chart.draw()

    def create_static_row(self, m, sigma, lmbd):
        static_row = []
        n = 40
        N = n * 1000
        if self.random_method_switch.currentText() == 'Метод серединных квадратов':
            x1 = random_values.random_mid_sqr(N)
            x2 = random_values.random_irrat_num(N)
        elif self.random_method_switch.currentText() == 'Метод иррационального числа':
            x1 = random_values.random_irrat_num(N)
            x2 = random_values.random_mid_sqr(N)
        else:
            x1 = random_values.random_cong(N)
            x2 = random_values.random_irrat_num(N)

        if self.distribution_switch.currentText() == 'Нормальное распределение':
            for i in range(N):
                static_row.append(random_values.normal_distribution(x1[i], x2[i], m, sigma))
        else:
            for i in range(N):
                static_row.append(random_values.exponential_distribution(x1[i], lmbd))
        return static_row

    def test_distribution(self):
        self.build_graph_gistogramm('create')
        #self.build_graph_acceptance_criterion('create')
        self.build_graph_acceptance_criterion('create')

        coord_x_ideal, coord_y_ideal = [], []
        m = float(self.random_test_m_cnt.text())
        sigma = float(self.random_test_sigma_cnt.text())
        lmbd = float(self.random_test_lambda_cnt.text())
        n = 40
        N = n * 1000
        static_row = self.create_static_row(m, sigma, lmbd)

        max = self.max_list_el(static_row)
        min = self.min_list_el(static_row)
        step = (max - min) / n # n

        dx = 0
        if self.distribution_switch.currentText() == 'Нормальное распределение':
            for i in range(n):
                coord_x_ideal.append(dx)
                coord_y_ideal.append(random_values.normal_ideal(dx, m, sigma))
                dx += step
            r = 2
        else:
            for i in range(n):
                coord_x_ideal.append(dx)
                coord_y_ideal.append(random_values.exponential_ideal(dx, lmbd))
                dx += step
            r = 3

        self.build_graph_gistogramm('build', coord_x_ideal, coord_y_ideal, 'r')

        intervals = []
        for i in range(n):
            intervals.append([min, min + step, 0])
            min += step

        for el in static_row:
            for interval in intervals:
                if el >= interval[0] and el <= interval[1]:
                    interval[2] += 1
                    break

        for interval in intervals:
            interval[2] /= N
            interval[2] /= step

        # Колмогоров
        coord_x, coord_y = [], []
        dx, dy = 0, 0
        Fcm = []
        for interval in intervals:
            coord_x.append(dx)
            dy += interval[2]
            Fcm.append(dy)
            coord_y.append(dy)
            coord_y.append(dy)
            dx += step
            coord_y.append(0)
            coord_x.append(dx)
            coord_x.append(dx)

        self.build_graph_acceptance_criterion('build', coord_x, coord_y, 'b')

        coord_x, coord_y = [], []
        dx, dy = 0, 0
        for i in range(n):
            if self.distribution_switch.currentText() == 'Экспоненциальное распределение':
                dy += coord_y_ideal[i] * 0.86
            else:
                dy += coord_y_ideal[i]
            Fcm[i] = abs(Fcm[i] - dy)
            coord_x.append(dx)
            coord_y.append(dy)
            dx += step

        D = self.max_list_el(Fcm)
        lmbd_Kolmogorov = D * sqrt(N)
        p = 1
        for k in range(-10000, 10000, 1):
            p -= (((-1) ** k) * exp(-2 * (k ** 2) * (lmbd_Kolmogorov ** 2)))
        self.kolmogorov_hypothesis.setText('Критерий согласия Колмогорова: %s' % (str((1 - p) * 100)) + '%')

        # Пирсон
        pirs = 0
        for i in range(n):
            raz = (intervals[i][2] * step) - coord_y_ideal[i]
            raz **= 2
            raz /= coord_y_ideal[i]
            pirs += raz
        pirs *= N
        pirs = float(str(sqrt(pirs)/10)[:str(sqrt(pirs)).find('e')])

        p_tabl = [99, 98, 95, 90, 80, 70, 50, 30]

        pirson_tabl = [[0.0002, 0.001, 0.004, 0.016, 0.064, 0.148, 0.455, 1.07],
        [0.020, 0.040, 0.103, 0.211, 0.446, 0.71, 1.39, 2.41],
        [0.115, 0.185, 0.352, 0.584, 1.01, 1.42, 2.37, 3.66],
        [0.297, 0.429, 0.711, 1.06, 1.65, 2.19, 3.36, 4.88],
        [0.554, 0.752, 1.15, 1.61, 2.34, 3.00, 4.35, 6.06],
        [0.872, 1.13, 1.64, 2.20, 3.07, 3.83, 5.35, 7.23],
        [1.24, 1.56, 2.17, 2.83, 3.82, 4.67, 6.35, 8.38],
        [1.65, 2.03, 2.73, 3.49, 4.59, 5.53, 7.34, 9.52],
        [2.09, 2.53, 3.33, 4.17, 5.38, 6.39, 8.34, 10.66],
        [2.56, 3.06, 3.94, 4.87, 6.18, 7.27, 9.34, 11.78],
        [3.05, 3.61, 4.57, 5.58, 6.99, 8.15, 10.34, 12.90],
        [3.57, 4.18, 5.23, 6.30, 7.81, 9.03, 11.34, 14.01],
        [4.11, 4.77, 5.89, 7.04, 8.63, 9.93, 12.34, 15.12],
        [4.66, 5.37, 6.57, 7.79, 9.47, 10.82, 13.34, 16.22],
        [5.23, 5.98, 7.26, 8.55, 10.31, 11.72, 14.34, 17.32],
        [5.81, 6.61, 7.96, 9.31, 11.15, 12.62, 15.34, 18.42],
        [6.41, 7.26, 8.67, 10.09, 12.00, 13.53, 16.34, 19.51],
        [7.01, 7.91, 9.39, 10.86, 12.86, 14.44, 17.34, 20.60],
        [7.63, 8.57, 10.12, 11.65, 13.72, 15.35, 18.34, 21.69],
        [8.26, 9.24, 10.85, 12.44, 14.58, 16.27, 19.34, 22.77],
        [8.90, 9.91, 11.59, 13.24, 15.44, 17.18, 20.34, 23.86],
        [9.54, 10.60, 12.34, 14.04, 16.31, 18.10, 21.34, 24.94],
        [10.20, 11.29, 13.09, 14.85, 17.19, 19.02, 22.34, 26.02],
        [10.86, 11.99, 13.85, 15.66, 18.06, 19.94, 23.34, 27.10],
        [11.52, 12.70, 14.61, 16.47, 18.94, 20.87, 24.34, 28.17],
        [12.20, 13.41, 15.38, 17.29, 19.82, 21.79, 25.34, 29.25],
        [12.88, 14.13, 16.15, 18.11, 20.70, 22.72, 26.34, 30.32],
        [13.56, 14.85, 16.93, 18.94, 21.59, 23.65, 27.34, 31.39],
        [14.26, 15.57, 17.71, 19.77, 22.48, 24.58, 28.34, 32.46],
        [14.95, 16.31, 18.49, 20.60, 23.36, 25.51, 29.34, 33.53],
        [15.66, 17.04, 19.28, 21.43, 24.26, 26.44, 30.34, 34.60],
        [16.36, 17.78, 20.07, 22.27, 25.15, 27.37, 31.34, 35.66],
        [17.07, 18.53, 20.87, 23.11, 26.04, 28.31, 32.34, 36.73],
        [17.79, 19.28, 21.66, 23.95, 26.94, 29.24, 33.34, 37.80],
        [18.51, 20.03, 22.47, 24.80, 27.84, 30.18, 34.34, 38.86],
        [19.23, 20.78, 23.27, 25.64, 28.73, 31.12, 35.34, 39.92],
        [19.96, 21.54, 24.07, 26.49, 29.64, 32.05, 36.34, 40.98],
        [20.69, 22.30, 24.88, 27.34, 30.54, 32.99, 37.34, 42.05],
        [21.43, 23.07, 25.70, 28.20, 31.44, 33.93, 38.34, 43.11],
        [22.16, 23.84, 26.51, 29.05, 32.34, 34.87, 39.34, 44.16],
        [22.91, 24.61, 27.33, 29.91, 33.25, 35.81, 40.34, 45.22],
        [23.65, 25.38, 28.14, 30.77, 34.16, 36.75, 41.34, 46.28],
        [24.40, 26.16, 28.96, 31.63, 35.07, 37.70, 42.34, 47.34],
        [25.15, 26.94, 29.79, 32.49, 35.97, 38.64, 43.34, 48.40],
        [25.90, 27.72, 30.61, 33.35, 36.88, 39.58, 44.34, 49.45],
        [26.66, 28.50, 31.44, 34.22, 37.80, 40.53, 45.34, 50.5],
        [27.42, 29.29, 32.27, 35.08, 38.71, 41.47, 46.34, 51.56]]

        for i in range(7, -1, -1):
            if pirs < pirson_tabl[n - r - 1][i]:
                break
        self.pirson_hypothesis.setText('Критерий согласия Пирсона: %s' % (str(p_tabl[i]) + '%'))

        # Корреляционный момент
        t = 1
        summ_1, summ_2, summ_3, summ_4, summ_5 = 0, 0, 0, 0, 0
        Kxy, Dx, Dy = 0, 0, 0
        minp = 0
        for i in range(N - t):
            summ_1 += static_row[i] * static_row[i + t]
            summ_2 += static_row[i]
            summ_3 += static_row[i + t]
            summ_4 += static_row[i] ** 2
            summ_5 += static_row[i + t] ** 2

        Kxy = (summ_1 / (N - t)) - (summ_2 * summ_3 / ((N - t) ** 2))
        Dx = (summ_4 / (N - t)) - ((summ_2 / (N - t)) ** 2)
        Dy = (summ_5 / (N - t)) - ((summ_3 / (N - t)) ** 2)
        pxy = abs(Kxy / sqrt(Dx * Dy))
        self.korel_moment.setText('Корелляционный метод: %s' % (str((1 - pxy) * 100) + '%'))

        # Гистограмма
        self.build_graph_acceptance_criterion('build', coord_x, coord_y, 'r')

        coord_x, coord_y = [], []

        dx = 0

        for i in range(n):
            coord_x.append(dx)
            coord_y.append(intervals[i][2])
            coord_y.append(intervals[i][2])
            dx += step
            coord_y.append(0)
            coord_x.append(dx)
            coord_x.append(dx)

        self.build_graph_gistogramm('build', coord_x, coord_y, 'b')

        sko = 0
        for i in range(len(coord_y_ideal)):
            sko += (coord_y_ideal[i] - intervals[i][2]) ** 2
        sko /= len(intervals)
        sko = sqrt(sko)

        self.SKO.setText('Среднеквадратичное отклонение: %s' % str(sko)[:8])

    def create_gist_coords(self, static_row):
        max = self.max_list_el(static_row)
        min = self.min_list_el(static_row)
        n = 40
        N = len(static_row)
        step = (max - min) / n

        intervals = []
        for i in range(n):
            intervals.append([min, min + step, 0])
            min += step

        for el in static_row:
            for interval in intervals:
                if el >= interval[0] and el <= interval[1]:
                    interval[2] += 1
                    break

        for interval in intervals:
            interval[2] /= N
            interval[2] /= step

        coord_x, coord_y = [], []
        dx, dy = 0, 0
        Fcm = []
        for interval in intervals:
            coord_x.append(dx)
            dy += interval[2]
            Fcm.append(dy)
            coord_y.append(dy)
            coord_y.append(dy)
            dx += step
            coord_y.append(0)
            coord_x.append(dx)
            coord_x.append(dx)

        out_1 = [coord_x, coord_y]

        dx = 0
        coord_x, coord_y = [], []
        for i in range(n):
            coord_x.append(dx)
            coord_y.append(intervals[i][2])
            coord_y.append(intervals[i][2])
            dx += step
            coord_y.append(0)
            coord_x.append(dx)
            coord_x.append(dx)

        out_2 = [coord_x, coord_y]

        return out_1, out_2, Fcm

    def smirnov(self, Nx, Ny):
        self.build_graph('create', self.figure_4, self.chart_4)
        out_1, out_2, Fcmx = self.create_gist_coords(Nx)
        coord_x, coord_y = out_2
        self.build_graph('build', self.figure_4, self.chart_4, coord_x, coord_y, 'b')
        self.build_graph('create', self.figure_5, self.chart_5)
        coord_x, coord_y = out_1
        self.build_graph('build', self.figure_5, self.chart_5, coord_x, coord_y, 'b')

        self.build_graph('create', self.figure_6, self.chart_6)
        out_1, out_2, Fcmy = self.create_gist_coords(Ny)
        coord_x, coord_y = out_2
        self.build_graph('build', self.figure_6, self.chart_6, coord_x, coord_y, 'b')
        self.build_graph('create', self.figure_7, self.chart_7)
        coord_x, coord_y = out_1
        self.build_graph('build', self.figure_7, self.chart_7, coord_x, coord_y, 'b')

        D = 0
        for i in range(len(Fcmx)):
            raz = abs(Fcmx[i] - Fcmy[i])
            if raz > D:
                D = raz

        p = 0
        p_list = [0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 0.95, 0.98, 0.99]
        for i in range(len(p_list)):
            s = sqrt(abs(log(1 - p_list[i]) * ((1 / len(Nx)) + (1 / len(Ny))) / 2)) * 100
            if D <= s:
                p = p_list[i]

        return p

    def student(self, Nx, Ny):
        mx, my = 0, 0
        lx = len(Nx)
        ly = len(Ny)
        for i in range(lx):
            mx += Nx[i]
        for i in range(ly):
            my += Ny[i]
        mx /= lx
        my /= ly
        Dx, Dy = 0, 0
        for i in range(lx):
            Dx += ((Nx[i] - mx) ** 2) / (lx - 1)
        for i in range(ly):
            Dy += ((Ny[i] - my) ** 2) / (ly - 1)

        D = ((lx - 1) * Dx + (ly - 1) * Dy) / (lx + ly - 2)
        t = sqrt((((mx - my) ** 2) * lx * ly) / (D * (lx + ly)))

        stud_coefs = [[0.126, 0.1], [0.253, 0.2], [0.385, 0.3], [0.524, 0.4], [0.674, 0.5],
                      [0.842, 0.6] ,[1.036, 0.7], [1.282, 0.8], [1.645, 0.9], [1.960, 0.95],
                      [2.330, 0.98], [2.580, 0.99], [3.290, 0.999]]

        for i in range(len(stud_coefs)):
            if t <= stud_coefs[i][0]:
                p = stud_coefs[i][1]
        return p

    def fisher(self, Nx, Ny):
        s = 0
        for i in range(0, len(Nx)):
            s += Nx[i]
        mx = (1 / len(Nx)) * s

        dx = 0
        for i in range(0, len(Nx)):
            dx += ((Nx[i] - mx) ** 2) / (len(Nx) - 1)

        s = 0
        for i in range(0, len(Ny)):
            s += Ny[i]
        my = (1 / len(Ny)) * s

        dy = 0
        for i in range(0, len(Ny)):
            dy += ((Ny[i] - my) ** 2) / (len(Ny) - 1)

        if dx > dy:
            F = dx / dy
        else:
            F = dy / dx

        F_tabl = 1.09
        if F <= F_tabl:
            return 'Cтатистическая гипотеза верна с вероятностью не менее, чем 95%'
        else:
            return 'Cтатистическая гипотеза верна с вероятностью менее, чем 95%'

    def array_for_disturbing(self):
        self.read_start_values()
        My = (self.sln_task_tm_rnd_max + self.sln_task_tm_rnd_min) / 2
        m, sigma = 0, 1
        N = 3000
        x11 = random_values.random_cong(N + 4)
        x12 = random_values.random_cong(N + 4)
        C0, C1, C2, C3 = 4.442, 0.496, 0.003, 4.946
        Q1 = []
        Q1 = [random_values.normal_distribution(x11[i], x12[i], m, sigma) for i in range(4)]
        Y1 = []
        for i in range(4, N + 4):
            Y1.append(C0 * Q1[i - 4] + C1 * Q1[i - 3] + C2 * Q1[i - 2] + C3 * Q1[i - 1] + My)
            Q1.append(random_values.normal_distribution(x11[i], x12[i], m, sigma))
        self.Y_work = [int(Y1[i]) + 1 for i in range(len(Y1))]

    def disturbing_proc_generate(self):
        self.read_start_values()
        a = ((self.ymax - self.ymin) / 6) ** 2
        b = -log(0.05) / 3
        func_K = lambda a, b, t: a * exp(-b * 3*t)
        self.build_graph('create', self.figure_3, self.chart_3)
        coord_x, coord_y = [], []
        i = 0
        while i <= 5:
            coord_x.append(i)
            coord_y.append(func_K(a, b, i))
            i += 0.1
        self.build_graph('build', self.figure_3, self.chart_3, coord_x, coord_y, 'r')

        My = (self.ymax + self.ymin) / 2
        m, sigma = 0, 1
        N = 1000
        x1 = random_values.random_cong(N + 4)
        x2 = random_values.random_cong(N + 4)
        C0, C1, C2, C3 = 4.442, 0.496, 0.003, 4.946
        Q = []
        Q = [random_values.normal_distribution(x1[i], x2[i], m, sigma) for i in range(4)]
        Y = []

        for i in range(4, N + 4):
            Y.append(C0 * Q[i - 4] + C1 * Q[i - 3] + C2 * Q[i - 2] + C3 * Q[i - 1] + My)
            Q.append(random_values.normal_distribution(x1[i], x2[i], m, sigma))

        Nx, Ny = [], []
        for i in range(N):
            if i % 2 == 0:
                Nx.append(Y[i])
            else:
                Ny.append(Y[i])
        self.Y = [int(Y[i]) + 1 for i in range(len(Y))]
        self.stud_info_output.setText('Cтатистическая гипотеза верна с вероятностью не менее, чем %s' % str(self.student(Nx, Ny) * 100) + '%')
        self.smirnov_info_output.setText('Cтатистическая гипотеза верна с вероятностью не менее, чем %s' % str(self.smirnov(Nx, Ny) * 100) + '%')
        self.fisher_info_output.setText(self.fisher(Nx, Ny))

    def next_time(self):
        if self.evm1.state != 'Free':
                self.evm1.next_time()
        if self.evm2.state != 'Free':
                self.evm2.next_time()
                # если ЭВМ2 освободилась, то освобождаем ЭВМ3
                if self.evm2.state == 'Free':
                    self.evm3.state = 'Free'

        if self.evm3.state != 'Free':
                self.evm3.next_time()
                # если ЭВМ3 освободилась, то освобождаем ЭВМ2
                if self.evm3.state == 'Free':
                    self.evm2.state = 'Free'

        # уменьшение времени до новой задачи
        self.tm_be4_newtask -= 1
        # увеличение реального времени
        self.tm_cnt += 1

        self.table_len += len(self.tasks_evm1) + len(self.tasks_ozu)+ len(self.evms)
        self.tableWidget.setRowCount(self.table_len)
        # запись алгоритма в файл + вывод в окно
        for evm in self.evms:
            self.file.write(str(self.tm_cnt) +' '+ evm.name +' '+ evm.state + '\n')
            self.tableWidget.setItem(self.k, 0, QTableWidgetItem(str(self.tm_cnt)))
            self.tableWidget.setItem(self.k, 1, QTableWidgetItem(evm.name))
            self.tableWidget.setItem(self.k, 2, QTableWidgetItem(evm.state))
            self.k += 1
        for task in self.tasks_evm1:
            self.file.write(str(self.tm_cnt) + ' ' + task.name +' '+ task.state + '\n')
            self.tableWidget.setItem(self.k, 0, QTableWidgetItem(str(self.tm_cnt)))
            self.tableWidget.setItem(self.k, 1, QTableWidgetItem(task.name))
            self.tableWidget.setItem(self.k, 2, QTableWidgetItem(task.state))
            self.k += 1
        for task in self.tasks_ozu:
            self.file.write(str(self.tm_cnt) + ' ' + task.name +' '+ task.state + '\n')
            self.tableWidget.setItem(self.k, 0, QTableWidgetItem(str(self.tm_cnt)))
            self.tableWidget.setItem(self.k, 1, QTableWidgetItem(task.name))
            self.tableWidget.setItem(self.k, 2, QTableWidgetItem(task.state))
            self.k += 1
        # сбор данных для анализа показателей работы
        # загруженность ЭВМ1
        if self.evm1.state.split()[0] == 'Work' and self.evm1.state.split()[2] == 'Task':
            self.tm_work_evm1 +=1
        # загруженность ЭВМ23
        if self.evm2.state.split()[0] == 'Work' and self.evm3.state.split()[0] == 'Work':
            if self.evm2.state.split()[2] == 'Task' and self.evm3.state.split()[2] == 'Task':
                self.tm_work_evm23 += 1
        # загруженность ОЗУ
        self.workload_ozu += (len(self.tasks_ozu)/self.size_ozu)

        # вывод показателей
        if self.tm_cnt == self.global_tm:
            self.table_len += 1
            self.tableWidget.setItem(self.k, 0, QTableWidgetItem('СЗ ЭВМ23'))
            self.workload_evm23 = (self.tm_work_evm23/self.global_tm)*100
            self.tableWidget.setItem(self.k, 1, QTableWidgetItem(str(round(self.workload_evm23,2))))
            self.tableWidget.setItem(self.k, 2, QTableWidgetItem(str(self.tm_work_evm23)))
            self.statistics_workload_evm23.setText(
                'Степень загруженности ЭВМ2,3: %s' % (str(round(self.workload_evm23,2)) +'%'))
            # степень загрузки ОЗУ
            self.workload_ozu /= self.global_tm
            self.workload_ozu *= 100
            self.statistics_workload_ozu.setText(
                'Степень загруженности ОЗУ: %s' % (str(round(self.workload_ozu, 2)) + '%'))
            # степень загрузки ЭВМ1
            self.workload_evm1 = (self.tm_work_evm1 / self.global_tm) * 100
            # вероятность удаления задания
            self.p_del_task /= self.last_task
            self.p_del_task *= 100
            self.statistics_p_del_task.setText(
                'Вероятность уничтожения задания: %s' % (str(round(self.p_del_task, 2)) + '%'))
        # Вывод показателей в Excel
            #self.import_excel()

    def import_excel(self):
        # Вывод показателей в Excel
        # Статистика одного экспримента
        '''wb = openpyxl.load_workbook('./Detail.xlsx')
        sheet = wb['Sheet1']
        sheet['A1'], sheet['B1'], sheet['C1'], sheet['D1'], sheet['E1'] = '№', 'СЗН', 'СЗ23', 'ВПН', 'КД'
        sheet['F1'],sheet['G1'], sheet['H1'], sheet['I1'], = 'СЗ1', 't1', 't2', 't23'
        sheet['J1'], sheet['K1'], sheet['L1'], sheet['L2'] = 'Дельта', 'КДД', 'N', self.size_ozu
        sheet['M1'], sheet['N1'], sheet['O1'] = 'm1', 'sigma1', 'lambda'
        sheet['M2'], sheet['N2'], sheet['O2'] = self.evm1_m, self.evm1_sigma, self.tm_be4_newtask_lambda
        sheet['M4'], sheet['N4'] = 'm2', 'sigma2'
        sheet['M5'], sheet['N5'] = self.evm2_m, self.evm2_sigma
        sheet['M7'], sheet['N7'] = 'm3', 'sigma3'
        sheet['M8'], sheet['N8'] = self.evm3_m, self.evm3_sigma
        sheet['A'+str(random_values.count_exp+1)] = random_values.count_exp
        sheet['B'+str(random_values.count_exp+1)] = round(self.workload_ozu, 2)
        sheet['C'+str(random_values.count_exp+1)] = round(self.workload_evm23, 2)
        sheet['D' + str(random_values.count_exp + 1)] = round(self.p_del_task, 2)
        sheet['E' + str(random_values.count_exp + 1)] = len(self.cmplt_task)
        sheet['F' + str(random_values.count_exp + 1)] = round(self.workload_evm1)
        sheet['G' + str(random_values.count_exp + 1)] = round(self.tm_between_newtasks, 2)
        sheet['H' + str(random_values.count_exp + 1)] = round(self.time_be4_cmplt_evm1, 2)
        sheet['I' + str(random_values.count_exp + 1)] = round(min(self.time_be4_cmplt_evm2, self.time_be4_cmplt_evm3), 2)
        sheet['J' + str(random_values.count_exp + 1)] = round(self.time_background_task, 2)
        sheet['K' + str(random_values.count_exp + 1)] = len(self.cmplt_b_task)
        wb.save('./Detail.xlsx')'''
        # Оптимизация
        wb = openpyxl.load_workbook('./Optim.xlsx')
        sheet = wb['Sheet1']
        sheet['A1'], sheet['B1'], sheet['C1'],  sheet['D1'] = '№', 'm3', 'm4', 'N'
        sheet['F1'], sheet['G1'], sheet['H1'],  = 'СЗ 23',  'ВУЗ', 'СЗ ОЗУ'
        sheet['A' + str(random_values.count_exp + 1)] = random_values.count_exp
        sheet['B' + str(random_values.count_exp + 1)] = self.evm3_m
        sheet['C' + str(random_values.count_exp + 1)] = self.evm2_m
        sheet['D' + str(random_values.count_exp + 1)] = self.size_ozu
        sheet['F' + str(random_values.count_exp + 1)] = round(self.workload_evm23, 2)
        sheet['G' + str(random_values.count_exp + 1)] = round(self.p_del_task, 2)
        sheet['H' + str(random_values.count_exp + 1)] = round(self.workload_ozu, 2)
        wb.save('./Optim.xlsx')
        # Протокол эксель не активен, т к низкое быстродействие записи
        '''wb = openpyxl.load_workbook('./Protocol.xlsx')
        sheet = wb['Sheet1']
        sheet['A1'], sheet['B1'], sheet['C1'], sheet['D1'], sheet['E1'] = 'Время', 'ЭВМ1', 'ЭВМ2', 'ЭВМ3', 'ОЗУ'
        sheet['G1'], sheet['H1'], sheet['I1'], = 't1', 't2', 't23'
        sheet['G2'], sheet['H2'], sheet['I2'], =  round(self.tm_between_newtasks, 2), round(self.time_be4_cmplt_evm1, 2), round(min(self.time_be4_cmplt_evm2, self.time_be4_cmplt_evm3), 2)
        sheet['J1'], sheet['K1'],  = 'Дельта', 'N',
        sheet['J2'], sheet['K2'], = round(self.time_background_task, 2), self.size_ozu
        sheet['M1'], sheet['N1'], sheet['O1'] = 'm1', 'sigma1', 'lambda'
        sheet['M2'], sheet['N2'], sheet['O2'] = self.evm1_m, self.evm1_sigma, self.tm_be4_newtask_lambda
        sheet['M4'], sheet['N4'] = 'm2', 'sigma2'
        sheet['M5'], sheet['N5'] = self.evm2_m, self.evm2_sigma
        sheet['M7'], sheet['N7'] = 'm3', 'sigma3'
        sheet['M8'], sheet['N8'] = self.evm3_m, self.evm3_sigma
        sheet['A' + str(self.tm_cnt)] = self.tm_cnt
        sheet['B' + str(self.tm_cnt)] = self.evm1.state
        sheet['C' + str(self.tm_cnt)] = self.evm2.state
        sheet['D' + str(self.tm_cnt)] = self.evm3.state

        if len(self.tasks_ozu) == 0:
            sheet['E' + str(self.tm_cnt)] = 'Free'
        if len(self.tasks_ozu) == 1:
            sheet['E' + str(self.tm_cnt)] = str(self.tasks_ozu[0].name)
        if len(self.tasks_ozu) == 2:
            sheet['E' + str(self.tm_cnt)] = str(self.tasks_ozu[0].name + '; '+ self.tasks_ozu[1].name)
        if self.size_ozu > 2:
            sheet['E' + str(self.tm_cnt)] = str(len(self.tasks_ozu) + '/ ' + self.size_ozu)

        #sheet['E' + str(self.tm_cnt)] = str(self.tasks_ozu[0]+'; '+self.tasks_ozu[1])

        wb.save('./Protocol.xlsx')'''

    def find_newname4task(self, type_task, num):
        if type_task == 'task':
            self.last_task = num
            return 'Task %s' % num
        if type_task == 'b_task':
            return 'B_task %s' % num

    def del_tasks(self):
        i = 0
        # Цикл для работы с очередью фоновых заданий
        while i != len(self.background_task):
            if self.background_task[i].state == 'Complete':
                    # добавление в список обработанных задач
                    self.cmplt_b_task.append(self.background_task[i])
                    # удаление этого задания из очереди к ЭВМ23
                    del self.background_task[i]
                    # Добавление в очередь нового фонового задания
                    self.b_num += 1
                    self.background_task.append(
                        Task(self.find_newname4task('b_task', self.b_num), self.time_be4_cmplt_evm1, self.time_background_task,
                             self.time_background_task))
            else:
                i += 1
        # Цикл для работы с ОЗУ и перемещением заявок в итоговой список
        i = 0
        while i != len(self.tasks_ozu):
            if self.tasks_ozu[i].state == 'Proc':
                # добавление в список обработанных заданий
                self.cmplt_task.append(self.tasks_ozu[i])
                # удаление этого задания из ОЗУ
                del self.tasks_ozu[i]
            else:
                i += 1
        # Цикл для работы с очередью перед ЭВМ1 и перемещением заявок в ОЗУ
        i = 0
        while i != len(self.tasks_evm1):
            if self.tasks_evm1[i].state == 'Complete':
                # проверка наличия свободного места в ОЗУ
                if len(self.tasks_ozu) < self.size_ozu:
                    # Увеличение времени до завершения обработки ЭВМ1 для избежания багов
                    self.tasks_evm1[i].time_be4_cmplt_evm1 = 1
                    # Изменение статуса задания на ожидание
                    self.tasks_evm1[i].set_wait_state()
                    # Изменение ЭВМ которые обрабатывают задание
                    self.tasks_evm1[i].name_evm = 23
                    # добавление в ОЗУ к ЭВМ2,3
                    self.tasks_ozu.append(self.tasks_evm1[i])
                    # удаление этого задания из очереди к ЭВМ1
                    del self.tasks_evm1[i]
                else:
                    del self.tasks_evm1[i]
                    self.p_del_task += 1
            else:
                i += 1

    def computing_device_modeling(self):
        random_values.count_exp += 1  # количество опытов моделирования
        self.read_start_values()
        self.b_num = 1
        self.num = 0
        self.k = 0
        self.table_len = 0
        self.memory = []
        if self.model_type_switch.currentText() == 'Детерминированная модель':
            model_type = 'det'
        elif self.model_type_switch.currentText() == 'Имитационная модель':
            model_type = 'imit'
        self.tm_cnt = 0
        self.tm_be4_newtask = 0
        self.evm1 = Evm(model_type, 1)
        self.evm2 = Evm(model_type, 2)
        self.evm3 = Evm(model_type, 3)
        self.evms = []  # список для ЭВМ
        self.evms.append(self.evm1)
        self.evms.append(self.evm2)
        self.evms.append(self.evm3)
        self.x1 = []
        self.x2 = []
        self.tasks_evm1 = []  # очередь для эвм 1
        self.tasks_ozu = []  # очередь для эвм 2,3
        self.background_task = []  # очередь для эвм 2,3 для фоновой задачи
        self.cmplt_task = []  # список обработанных заданий
        self.cmplt_b_task = []  # список обработанных фоновых заданий
        self.tm_work_evm23 = 0  # время работы ЭВМ23 над основными заданиями
        self.workload_evm23 = 0  # степень загрузки ЭВМ23
        self.tm_work_evm1 = 0  # время работы ЭВМ1 над заданиями
        self.workload_evm1 = 0   # степень загрузки ЭВМ1
        self.workload_ozu = 0  # степень загрузки ОЗУ
        self.p_del_task = 0    # вероятность уничтожения задания
        self.last_task = 0     # номер последнего задания для нахождения вероятности
        value_work = [[0] * 165 for i in range(3)]   # двумерный массив для паретто, критерии
        parameters = [[0] * 165 for i in range(2)]  # двумерный массив для паретто, параметры
        time_background_task_cnt = 0
        if model_type == 'imit':
            # создаем массив для генарации возмущения
            self.array_for_disturbing()
            # генерируем массивы из ПСЧ
            self.x1 = random_values.random_cong(1000)
            self.x2 = random_values.random_irrat_num(1000)
            # Определяем случайные значения параметров системы
            self.tm_between_newtasks = int(abs(random_values.exponential_distribution(choice(self.x1),
                                                                                      self.tm_be4_newtask_lambda)))
            self.time_be4_cmplt_evm1 = int(abs(random_values.normal_distribution(
                choice(self.x1), choice(self.x2), self.evm1_m, self.evm1_sigma)))
            self.time_be4_cmplt_evm2 = int(abs(random_values.normal_distribution(
                choice(self.x1), choice(self.x2), self.evm2_m, self.evm2_sigma)))
            self.time_be4_cmplt_evm3 = int(abs(random_values.normal_distribution(
                choice(self.x1), choice(self.x2), self.evm3_m, self.evm3_sigma)))

        self.file = open('info.txt', 'w')
        # реальное время меньше модельного?
        while self.tm_cnt < self.global_tm:
            # Пришло ли время задания?
            if self.tm_be4_newtask <= 0:
                # Время нового задание равно периоду поступления заданий t1
                self.tm_be4_newtask = self.tm_between_newtasks
                # Добавление в очередь к ЭВМ1 нового задания
                if len(self.tasks_evm1) < 16:
                    self.num += 1
                    self.tasks_evm1.append(
                        Task(self.find_newname4task('task', self.num), self.time_be4_cmplt_evm1,
                             self.time_be4_cmplt_evm2,
                             self.time_be4_cmplt_evm3))
            # Добавление в очередь фонового задания
            if len(self.background_task) < 5:
                self.background_task.append(Task(self.find_newname4task('b_task', self.b_num), self.time_be4_cmplt_evm1,
                                                 self.time_background_task, self.time_background_task))
            # Обработка задания ЭВМ1
            if self.evm1.state == 'Free':
                for task in self.tasks_evm1:
                    if task.state == 'Wait':
                        self.evm1.set_work_state(task)
                        break

            # ЭВМ1 обработала задание?
            # если да, то удаление этого задания из очереди к ЭВМ1 и добавление в ОЗУ к ЭВМ2,3
            self.del_tasks()

            # Обработка задания ЭВМ2 и ЭВМ3
            if self.evm2.state == 'Free' and self.evm3.state == 'Free':
                # Если ЭВМ2 и ЭВМ3 свободны и ОЗУ пуста, то решается фоновая задача
                if len(self.tasks_ozu) == 0:
                    # задание возмущения
                    if model_type == 'imit':
                        self.time_background_task = int(self.Y_work[time_background_task_cnt])
                        time_background_task_cnt += 1
                    for background_task in self.background_task:
                        if background_task.state == 'Wait':
                            self.evm2.set_work_state(background_task)
                            self.evm3.set_work_state(background_task)
                            break
                for task in self.tasks_ozu:
                    if task.state == 'Wait':
                        self.evm2.set_work_state(task)
                        self.evm3.set_work_state(task)
                        break

            # увеличение реального времени
            self.next_time()

        #self.pareto(parameters, value_work)

        self.file.close()

    # парето, заполняем массив из экселя
    def pareto(self, parameters, value_work):
        wb = openpyxl.load_workbook('./Optim.xlsx')
        sheet = wb['Sheet1']
        j = 0
        while j < 165:
            value_work[0][j] = sheet['F' + str(j + 2)]
            value_work[1][j] = sheet['G' + str(j + 2)]
            value_work[2][j] = sheet['H' + str(j + 2)]
            parameters[0][j] = sheet['B' + str(j + 2)]
            parameters[1][j] = sheet['D' + str(j + 2)]
            j += 1
        # оптимизация
        j = 0
        while j < len(value_work[0]) - 1:
            k = 0
            while k < len(value_work[0]):
                if k != j:
                    if (value_work[0][k].value >= value_work[0][j].value) and (
                            value_work[1][k].value <= value_work[1][j].value) and (
                            value_work[2][k].value <= value_work[2][j].value) and (
                            parameters[1][k].value <= parameters[1][j].value) or (value_work[0][j].value < 90) or (
                            value_work[1][j].value > 0):
                        del value_work[0][j]
                        del value_work[1][j]
                        del value_work[2][j]
                        del parameters[0][j]
                        del parameters[1][j]
                        j -= 1
                        break
                k += 1
            j += 1
        wb.save('./Optim.xlsx')
        # вывод результатов оптимизации
        wb = openpyxl.load_workbook('./optim_result.xlsx')
        sheet = wb['Sheet1']
        sheet['A1'], sheet['B1'], sheet['C1'] = '№', 'm3', 'N'
        sheet['F1'], sheet['G1'], sheet['H1'], = 'СЗ 23', 'ВУЗ', 'СЗ ОЗУ'
        j = 0
        while j < len(value_work[0]):
            sheet['A' + str(j + 2)] = j + 1
            sheet['B' + str(j + 2)] = parameters[0][j].value
            sheet['C' + str(j + 2)] = parameters[1][j].value
            sheet['F' + str(j + 2)] = value_work[0][j].value
            sheet['G' + str(j + 2)] = value_work[1][j].value
            sheet['H' + str(j + 2)] = value_work[2][j].value
            j += 1
        wb.save('./optim_result.xlsx')

if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = MainWindow()
    sys.exit(app.exec_())
